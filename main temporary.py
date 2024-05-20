import pandas as pd
import os
import logging
from tkinter import *
from tkinter import messagebox
from datetime import datetime, timedelta
import time
import shutil
import glob # For file pattern matching
import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.styles.numbers import BUILTIN_FORMATS
import win32com.client # For detecting if file is open and then closing it
import subprocess # For file opening

# ---------------------------- BACKEND SETUP ------------------------------- #
# Logging (enhanced for more informative messages)
logging.basicConfig(
    level=logging.DEBUG, 
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Active Directory
# Get the absolute path of the script and use it as active directory
active_directory = os.path.dirname(os.path.abspath(__file__))
os.chdir(active_directory)

# ---------------------------- FILE MANAGEMENT ----------------------------- #
trading_journal_path = "Output/trading_journal_live.xlsx"
trading_dashboard_path = "Output/trading_dashboard_live.xlsx"
tradersync_export_path = "Imports/trade_data.csv"

def create_backup(file_path):
    """Creates a timestamped backup and deletes previous backups of the given file.
    
    Args:
        file_path (string): file path for the file that will be handled.
    """
    try:
        # Backup file creation
        file_dir, file_name = os.path.split(file_path)
        backup_name = os.path.join(file_dir, f"{os.path.splitext(file_name)[0]}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
        shutil.copy(file_path, backup_name)

        # Delete previous backups of the same file
        backup_pattern = os.path.join(file_dir, f"{os.path.splitext(file_name)[0]}_*.xlsx")
        for old_backup in glob.glob(backup_pattern):
            if old_backup != backup_name:
                try:
                    os.remove(old_backup)
                    logger.info(f"Deleted previous backup: {old_backup}")
                except Exception as e:
                    logger.warning(f"Failed to delete backup: {old_backup}. Error: {e}")

        logger.info(f"Backup created: {backup_name}")
        logger.info(f"Backup created at: {os.path.abspath(backup_name)}")
    except Exception as e:
        logger.error(f"Backup or deletion failed: {e}")
        messagebox.showerror(title="Error", message=f"An error occurred during backup: {e}")

def open_file(file_path):
    """Open the file using the default application.
    
    Args:
        file_path (string): file path for the file that will be handled.
    """
    try:
        # Open the file with the default associated application
        subprocess.Popen(["start", "", file_path], shell=True)
        logger.info(f"open_file: Opened the file: {file_path}")
    except Exception as e:
        logger.error(f"open_file: An error occurred: {e}")
        messagebox.showerror(title="Error", message=f"An error occurred while opening the file: {e}")
        
def close_open_file(file_path):
    """Check if the specified file is open and close it if necessary.
    
    Args:
        file_path (string): file path for the file that will be handled.
    """
    try:
        # Normalize and get the absolute file path for the operating system, with lower case drive letter
        normalized_file_path = os.path.abspath(os.path.normpath(file_path)).lower()

        # Create an instance of the Excel application
        excel = win32com.client.Dispatch("Excel.Application")

        # Debug: Log the normalized file path
        logger.debug(f"Normalized file path to close: {normalized_file_path}")

        # Extract the filename from the file path
        target_filename = os.path.basename(normalized_file_path)

        # Iterate through the open workbooks
        for workbook in excel.Workbooks:
            workbook_path = os.path.abspath(os.path.normpath(workbook.FullName)).lower()

            # Debug: Log the workbook path
            logger.debug(f"Open workbook path: {workbook_path}")

            # Extract the filename from the workbook path
            workbook_filename = os.path.basename(workbook_path)

            # Compare filenames to handle OneDrive paths and case differences
            if workbook_filename == target_filename:
                workbook.Close(SaveChanges=False)
                logger.info(f"close_open_file: Closed the open file: {normalized_file_path}")
                break
        else:
            logger.info(f"close_open_file: File is not open: {normalized_file_path}")

        # Ensure proper cleanup
        del excel
        time.sleep(1)  # Add a short delay to ensure the Excel process releases the file
    except Exception as e:
        logger.error(f"close_open_file: An error occurred: {e}")
        messagebox.showerror(title="Error", message=f"An error occurred while closing the file: {e}")

# ---------------------------- ERROR CHECKERS ------------------------------ #
def check_write_permission(directory):
    """Check if the user has writer permission to create file backup

    Args:
        directory (string): current directory where the code is being run from

    Returns:
        None: Uses return to stop the code from executing as no backup can be made
    """    
    if os.access(directory, os.W_OK):
        logger.info(f"Write permission exists for: {directory}")
        return True
    else:
        logger.error(f"No write permission for: {directory}")
        return False

# ---------------------------- CONSTANTS ----------------------------------- #
FONT_NAME = "Calibri"
FONT_SIZE = 11
WHITE = "#fcf7f9"
MARKET_TICKERS = ["SPY", "MES", "QQQ", "CONGLO", "IWM", "VIX"]  # Tickers that track overall markets
# ---------------------------- INPUT CHECKERS ------------------------------ #

# ---------------------------- WEEKLY TASKS -------------------------------- #
def weekly_tasks():
    """Runs the weekly tasks as defined in the instructions.
    These are activated via a button in the UI.\n
    - Closes open files so they can be properly handled\n
    - Creates backup of files that will be manipulated\n
    - Clears and creates the Weekly One Pager\n
    - Opens the file upon finishing processing

    Returns:
        None: Uses return to stop the code from executing as no backup can be made
    """    
    if not check_write_permission(active_directory):
        messagebox.showerror(title="Error", message="Insufficient permissions. Please check if you have write access to the output directory.")
        return  # Stop execution if no write permission

    close_open_file(trading_journal_path)  # Ensure the file is not open
    create_backup(trading_journal_path)    # Create backup before modifying
    clear_one_pager()
    weekly_journal_processor()
    open_file(trading_journal_path)        # Open the file after processing

def clear_one_pager():
    """Clears the Weekly One Pager sheet
    """    
    try:
        # Load the workbook 
        book = openpyxl.load_workbook(trading_journal_path)  
        
        # Select the sheet
        sheet = book['Weekly One Pager']

        # Clear all rows except the header
        sheet.delete_rows(2, sheet.max_row)  

        # Save the changes
        book.save(trading_journal_path) 

        logger.info("clear_one_pager: Weekly One Pager sheet has been cleared successfully.")
    except Exception as e:
        logger.error(f"clear_one_pager: An error occurred: {e}")

def weekly_journal_processor():
    """Function to identify all the Journal notes relevant to the current week, sorting them by date and priority, and then adding them to Weekly One Pager."""
    try:
        # Load the Journal sheet
        journal_df = pd.read_excel(trading_journal_path, sheet_name="Journal")

        # Ensure 'Date' column is in datetime format without time component
        journal_df['Date'] = pd.to_datetime(journal_df['Date']).dt.date

        # Get current date and calculate the start and end of the current week
        today = datetime.now().date()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)

        # Filter rows for the current week and where "Weekly One Pager" is not 'no'
        current_week_df = journal_df[
            (journal_df['Date'] >= start_of_week) &
            (journal_df['Date'] <= end_of_week) &
            (journal_df['Weekly One Pager'] != 'no')
        ]

        # Create a priority column for sorting
        current_week_df['Priority'] = current_week_df.apply(
            lambda row: (0 if row['Ticker'] in MARKET_TICKERS else (1 if row['Followup'] == 'Yes' else 2)), 
            axis=1
        )

        # Sort by Date (descending) and Priority
        current_week_df = current_week_df.sort_values(by=['Date', 'Priority'], ascending=[False, True])

        # Select required columns and make a copy
        selected_columns = current_week_df[['Date', 'Ticker', 'Comments', 'Key Support Level', 'Key Resistance Level', 'Weekly One Pager']].copy()

        # Rename columns in the copied DataFrame
        selected_columns.rename(columns={
            'Weekly One Pager': 'Game Plan'
        }, inplace=True)

        # Append to Weekly One Pager sheet
        with pd.ExcelWriter(trading_journal_path, mode='a', if_sheet_exists='overlay') as writer:
            workbook = writer.book #FIXME is this used?
            worksheet = writer.sheets["Weekly One Pager"]
            start_row = worksheet.max_row  # Find the next empty row
            selected_columns.to_excel(writer, sheet_name="Weekly One Pager", index=False, startrow=start_row, header=False)

        logger.info("weekly_journal_processor: Rows have been copied to the Weekly One Pager sheet successfully.")
        messagebox.showinfo(title="Success", message="Weekly One Pager Built")
    except Exception as e:
        # Log the specific error instead of just the exception object
        logger.error(f"weekly_journal_processor: An error occurred: {e}") 
        # Display error message to user
        messagebox.showerror(title="Error", message=f"An error occurred while processing the journal: {e}")

# ---------------------------- DAILY TASKS --------------------------------- #
def daily_tasks():
    """Runs the daily tasks as defined in the instructions.
    These are activated via a button in the UI.\n
    - Closes open files so they can be properly handled\n
    - Creates backup of files that will be manipulated\n
    - Imports the tradersync (TDSync) export from the Import folder\n
    - Processes the TDSync export to bring closed trades to the Trade Log\n
    - Open file after processing is finished
    #TODO
    """
    close_open_file(trading_journal_path)  # Ensure the file is not open
    close_open_file(tradersync_export_path) # Ensure the file is not open
    create_backup(trading_journal_path)    # Create backup before modifying    
    tradersync_import() # Copy Tradersync export into the TradersyncExport sheet
    process_tradersync_export() # Process the export into the trade log. Only process non OPEN entries.
    open_file(trading_journal_path) # Open the file after processing

import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.styles.numbers import BUILTIN_FORMATS

def tradersync_import():
    """Copies trade_data.csv content to the TraderSync Export sheet in the trading_journal_path without altering the header."""
    try:
        # Load the CSV file
        trade_data_df = pd.read_csv(tradersync_export_path)

        # Remove dollar signs ('$') from currency columns and convert them to float
        currency_columns = ['Entry Price', 'Exit Price', 'Return $', 'Avg Buy', 'Avg Sell', 
                            'Net Return', 'Commision', 'Strike', 'Cost', 'Fees', 
                            'Return Share', 'Best Exit $']
        trade_data_df[currency_columns] = trade_data_df[currency_columns].replace('[\$,]', '', regex=True).astype(float)

        # Load the Excel file
        book = openpyxl.load_workbook(trading_journal_path)
        sheet = book['TraderSync Export']

        # Clear the sheet except the header
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.value = None

        # Get the built-in currency format
        currency_format = BUILTIN_FORMATS[7]

        # Apply the currency format to specified columns
        for col_name in currency_columns:
            col_idx = trade_data_df.columns.get_loc(col_name) + 1
            for row_idx, value in enumerate(trade_data_df[col_name], start=2):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.number_format = currency_format

        # Write the combined data back to the sheet
        for idx, row in trade_data_df.iterrows():
            for col_idx, value in enumerate(row):
                sheet.cell(row=idx + 2, column=col_idx + 1, value=value)

        # Save the workbook
        book.save(trading_journal_path)

        logger.info("tradersync_import: Trade data copied to TraderSync Export sheet successfully.")
        messagebox.showinfo(title="Success", message="TraderSync import completed successfully.")
    except Exception as e:
        logger.error(f"tradersync_import: An error occurred: {e}")
        messagebox.showerror(title="Error", message=f"An error occurred while processing the TraderSync import: {e}")

def process_tradersync_export():
    """Reads data from TraderSync Export sheet and appends specific values to Trade Log sheet without altering the headers, only if Status is different from 'OPEN'. Finally, sorts the data by Close Date in descending order and adds a formula to Net Return %."""
    try:
        # Load the Excel file
        book = openpyxl.load_workbook(trading_journal_path)

        # Load the TraderSync Export and Trade Log sheets
        tradersync_export_sheet = book['TraderSync Export']
        trade_log_sheet = book['Trade Log']

        # Define the columns to be copied
        columns_to_copy = [
            'Status', 'Symbol', 'Size', 'Open Date', 'Close Date', 'Avg Buy', 
            'Avg Sell', 'Net Return', 'Type', 'MAE', 'MFE', 'Best Exit $', 'Best Exit %'
        ]

        # Get headers from the TraderSync Export sheet
        tradersync_headers = [cell.value for cell in tradersync_export_sheet[1]]

        # Create a dictionary to map column names to indices
        tradersync_col_index = {name: idx + 1 for idx, name in enumerate(tradersync_headers)}

        # Load existing data from Trade Log sheet into a DataFrame
        trade_log_headers = [cell.value for cell in trade_log_sheet[1]]
        existing_data = pd.DataFrame(trade_log_sheet.iter_rows(values_only=True, min_row=2), columns=trade_log_headers)

        # Get the maximum Trade ID from the existing data
        max_trade_id = existing_data['Trade ID'].dropna().astype(int).max() if not existing_data.empty else 0

        # Initialize the next Trade ID
        next_trade_id = max_trade_id + 1

        # Iterate over rows in the TraderSync Export sheet, starting from the second row (skip headers)
        new_data_rows = []
        for row in tradersync_export_sheet.iter_rows(min_row=2, values_only=True):
            # Check if the Status value is different from 'OPEN'
            status_col_index = tradersync_col_index.get('Status')
            if status_col_index and row[status_col_index - 1] != 'OPEN':
                # Create a list to hold the values to be appended
                row_data = [next_trade_id]  # Add the new Trade ID
                for col_name in columns_to_copy:
                    col_index = tradersync_col_index.get(col_name)
                    if col_index:
                        row_data.append(row[col_index - 1])
                    else:
                        row_data.append(None)  # Append None if the column name is not found
                new_data_rows.append(row_data)
                next_trade_id += 1  # Increment the Trade ID for the next entry

        # Create a DataFrame for the new data
        new_data_df = pd.DataFrame(new_data_rows, columns=['Trade ID'] + columns_to_copy)

        # Combine existing data with the new data
        combined_data = pd.concat([existing_data, new_data_df], ignore_index=True)

        # Sort the combined data by Close Date in descending order. Errors will handle stuff not being recognized as dates.
        combined_data['Close Date'] = pd.to_datetime(combined_data['Close Date'], errors='coerce')
        combined_data = combined_data.sort_values(by='Close Date', ascending=False).reset_index(drop=True)  # Reset index after sorting

        # Clear the Trade Log sheet except the header
        for row in trade_log_sheet.iter_rows(min_row=2, max_row=trade_log_sheet.max_row):
            for cell in row:
                cell.value = None

        # Get column indices for Trade ID, Avg Buy, Avg Sell, Net Return %, Net Return, MAE, MFE, Best Exit $, and Best Exit %
        trade_id_col_idx = trade_log_headers.index('Trade ID') + 1
        avg_buy_col_idx = trade_log_headers.index('Avg Buy') + 1
        avg_sell_col_idx = trade_log_headers.index('Avg Sell') + 1
        net_return_pct_col_idx = trade_log_headers.index('Net Return %') + 1
        net_return_col_idx = trade_log_headers.index('Net Return') + 1
        mae_col_idx = trade_log_headers.index('MAE') + 1
        mfe_col_idx = trade_log_headers.index('MFE') + 1
        best_exit_col_idx = trade_log_headers.index('Best Exit $') + 1
        best_exit_pct_col_idx = trade_log_headers.index('Best Exit %') + 1

        # Apply the built-in currency format to specified columns
        currency_format = BUILTIN_FORMATS[7]
        percentage_format = '0.00%'

        # Write the sorted data back to the Trade Log sheet and add the formula for Net Return %
        for idx, row in combined_data.iterrows():
            for col_idx, value in enumerate(row, start=1):
                cell = trade_log_sheet.cell(row=idx + 2, column=col_idx, value=value)
                if col_idx in [avg_buy_col_idx, avg_sell_col_idx, net_return_col_idx, mae_col_idx, mfe_col_idx, best_exit_col_idx]:
                    cell.number_format = currency_format
                elif col_idx == best_exit_pct_col_idx:
                    cell.number_format = percentage_format
                    cell.value = value / 100  # Divide by 100 to correctly show percentage
            
            # Set the formula for Net Return % in the current row
            net_return_pct_cell = trade_log_sheet.cell(row=idx + 2, column=net_return_pct_col_idx)
            net_return_pct_cell.value = f"=({trade_log_sheet.cell(row=idx + 2, column=avg_sell_col_idx).coordinate}-{trade_log_sheet.cell(row=idx + 2, column=avg_buy_col_idx).coordinate})/{trade_log_sheet.cell(row=idx + 2, column=avg_buy_col_idx).coordinate}"
            net_return_pct_cell.number_format = '0.00%'

        # Save the workbook
        book.save(trading_journal_path)

        logger.info("process_tradersync_export: Data from TraderSync Export sheet appended and sorted in Trade Log sheet successfully.")
        messagebox.showinfo(title="Success", message="Data processed, appended, and sorted in Trade Log successfully.")
    except Exception as e:
        logger.error(f"process_tradersync_export: An error occurred: {e}")
        messagebox.showerror(title="Error", message=f"An error occurred while processing the TraderSync export: {e}")

# ---------------------------- UI SETUP ------------------------------------ #
# Main window UI setup
main_window = Tk()
main_window.title("Trading Journal")
main_window.config(padx=50, pady=50, bg=WHITE)

# Action button that will run the weekly tasks
weekly_task_button = Button(text="Perform Weekly Tasks", command=weekly_tasks)
weekly_task_button.grid(column=0, row=0)

# Action button that will run the daily tasks
daily_task_button = Button(text="Perform Daily Tasks", command=daily_tasks)
daily_task_button.grid(column=0, row=1)

main_window.mainloop()
