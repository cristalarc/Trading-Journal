import pandas as pd
import os
import logging
from tkinter import *
from tkinter import messagebox
from datetime import datetime, timedelta
import time
import shutil
import glob # For file pattern matching
import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.styles.numbers import BUILTIN_FORMATS
import win32com.client # For detecting if file is open and then closing it
import subprocess # For file opening

# ---------------------------- BACKEND SETUP ------------------------------- #
# Logging (enhanced for more informative messages)
logging.basicConfig(
    level=logging.DEBUG, 
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Active Directory
# Get the absolute path of the script and use it as active directory
active_directory = os.path.dirname(os.path.abspath(__file__))
os.chdir(active_directory)

# ---------------------------- FILE MANAGEMENT ----------------------------- #
trading_journal_path = "Output/trading_journal_live.xlsx"
trading_dashboard_path = "Output/trading_dashboard_live.xlsx"
tradersync_export_path = "Imports/trade_data.csv"

def create_backup(file_path):
    """Creates a timestamped backup and deletes previous backups of the given file.
    
    Args:
        file_path (string): file path for the file that will be handled.
    """
    try:
        # Backup file creation
        file_dir, file_name = os.path.split(file_path)
        backup_name = os.path.join(file_dir, f"{os.path.splitext(file_name)[0]}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
        shutil.copy(file_path, backup_name)

        # Delete previous backups of the same file
        backup_pattern = os.path.join(file_dir, f"{os.path.splitext(file_name)[0]}_*.xlsx")
        for old_backup in glob.glob(backup_pattern):
            if old_backup != backup_name:
                try:
                    os.remove(old_backup)
                    logger.info(f"Deleted previous backup: {old_backup}")
                except Exception as e:
                    logger.warning(f"Failed to delete backup: {old_backup}. Error: {e}")

        logger.info(f"Backup created: {backup_name}")
        logger.info(f"Backup created at: {os.path.abspath(backup_name)}")
    except Exception as e:
        logger.error(f"Backup or deletion failed: {e}")
        messagebox.showerror(title="Error", message=f"An error occurred during backup: {e}")

def open_file(file_path):
    """Open the file using the default application.
    
    Args:
        file_path (string): file path for the file that will be handled.
    """
    try:
        # Open the file with the default associated application
        subprocess.Popen(["start", "", file_path], shell=True)
        logger.info(f"open_file: Opened the file: {file_path}")
    except Exception as e:
        logger.error(f"open_file: An error occurred: {e}")
        messagebox.showerror(title="Error", message=f"An error occurred while opening the file: {e}")
        
def close_open_file(file_path):
    """Check if the specified file is open and close it if necessary.
    
    Args:
        file_path (string): file path for the file that will be handled.
    """
    try:
        # Normalize and get the absolute file path for the operating system, with lower case drive letter
        normalized_file_path = os.path.abspath(os.path.normpath(file_path)).lower()

        # Create an instance of the Excel application
        excel = win32com.client.Dispatch("Excel.Application")

        # Debug: Log the normalized file path
        logger.debug(f"Normalized file path to close: {normalized_file_path}")

        # Extract the filename from the file path
        target_filename = os.path.basename(normalized_file_path)

        # Iterate through the open workbooks
        for workbook in excel.Workbooks:
            workbook_path = os.path.abspath(os.path.normpath(workbook.FullName)).lower()

            # Debug: Log the workbook path
            logger.debug(f"Open workbook path: {workbook_path}")

            # Extract the filename from the workbook path
            workbook_filename = os.path.basename(workbook_path)

            # Compare filenames to handle OneDrive paths and case differences
            if workbook_filename == target_filename:
                workbook.Close(SaveChanges=False)
                logger.info(f"close_open_file: Closed the open file: {normalized_file_path}")
                break
        else:
            logger.info(f"close_open_file: File is not open: {normalized_file_path}")

        # Ensure proper cleanup
        del excel
        time.sleep(1)  # Add a short delay to ensure the Excel process releases the file
    except Exception as e:
        logger.error(f"close_open_file: An error occurred: {e}")
        messagebox.showerror(title="Error", message=f"An error occurred while closing the file: {e}")

def excel_cell_formater(format):
    """Applies an excel format to the requested cell.
    You can assign with cell.number_format = call this function.

    Args:
        format (string: %,USD, date, etc): list of available formats that can be implemented.
            %: Excel percentage cell format
            USD: Excel currency $ format
            date: Excel mm/dd/yyyy date format
    """    
    # Apply the built-in currency format to specified columns
    currency_USD_format = BUILTIN_FORMATS[7]
    percentage_format = '0.00%'
    date_format = BUILTIN_FORMATS[14]
    if format == "USD":
        return currency_USD_format
    elif format == "%":
        return percentage_format
    elif format == "date":
        return date_format

# ---------------------------- ERROR CHECKERS ------------------------------ #
def check_write_permission(directory):
    """Check if the user has writer permission to create file backup

    Args:
        directory (string): current directory where the code is being run from

    Returns:
        None: Uses return to stop the code from executing as no backup can be made
    """    
    if os.access(directory, os.W_OK):
        logger.info(f"Write permission exists for: {directory}")
        return True
    else:
        logger.error(f"No write permission for: {directory}")
        return False

# ---------------------------- CONSTANTS ----------------------------------- #
FONT_NAME = "Calibri"
FONT_SIZE = 11
WHITE = "#fcf7f9"
MARKET_SYMBOLS = ["SPY", "MES", "QQQ", "CONGLO", "IWM", "VIX"]  # Tickers/Symbols that track overall markets
# ---------------------------- INPUT CHECKERS ------------------------------ #

# ---------------------------- WEEKLY TASKS -------------------------------- #
def process_one_pager_task():
    """Runs the weekly tasks as defined in the instructions.
    These are activated via a button in the UI.\n
    - Closes open files so they can be properly handled\n
    - Creates backup of files that will be manipulated\n
    - Clears and creates the Weekly One Pager\n
    - Opens the file upon finishing processing

    Returns:
        None: Uses return to stop the code from executing as no backup can be made
    """    
    if not check_write_permission(active_directory):
        messagebox.showerror(title="Error", message="Insufficient permissions. Please check if you have write access to the output directory.")
        return  # Stop execution if no write permission

    close_open_file(trading_journal_path)  # Ensure the file is not open
    create_backup(trading_journal_path)    # Create backup before modifying
    clear_one_pager()
    journal_processor()
    open_file(trading_journal_path)        # Open the file after processing

def clear_one_pager():
    """Clears the Weekly One Pager sheet
    """    
    try:
        # Load the workbook 
        book = openpyxl.load_workbook(trading_journal_path)  
        
        # Select the sheet
        sheet = book['Weekly One Pager']

        # Save data validations
        data_validations = sheet.data_validations

        # Clear all rows except the header
        sheet.delete_rows(2, sheet.max_row)  

        # Reapply data validations
        sheet.data_validations = data_validations

        # Save the changes
        book.save(trading_journal_path) 

        logger.info("clear_one_pager: Weekly One Pager sheet has been cleared successfully.")
    except Exception as e:
        logger.error(f"clear_one_pager: An error occurred: {e}")

def journal_processor():
    """Function to identify all the Journal notes relevant to the current week, sorting them by date and priority, and then adding them to Weekly One Pager."""
    try:
        # Load the Journal sheet
        journal_df = pd.read_excel(trading_journal_path, sheet_name="Journal")

        # Ensure 'Date' column is in datetime format without time component
        journal_df['Date'] = pd.to_datetime(journal_df['Date']).dt.date

        # Get current date and calculate the start and end of the current week
        today = datetime.now().date()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)

        # Filter rows for the current week and where "Weekly One Pager" is not 'no'
        current_week_df = journal_df[
            (journal_df['Date'] >= start_of_week) &
            (journal_df['Date'] <= end_of_week) &
            (journal_df['Weekly One Pager'] != 'no')
        ].copy()  # Using .copy() to avoid SettingWithCopyWarning

        # Create a priority column for sorting
        current_week_df.loc[:, 'Priority'] = current_week_df.apply(
            lambda row: (0 if row['Symbol'] in MARKET_SYMBOLS else (1 if row['Followup'] == 'Yes' else 2)), 
            axis=1
        )

        # Sort by Date (descending) and Priority
        current_week_df = current_week_df.sort_values(by=['Date', 'Priority'], ascending=[False, True])
        #FIXME Change this to just priority if preferable

        # Select required columns and make a copy
        selected_columns = current_week_df[['Date', 'Symbol', 'Comments', 'Key Support Level', 'Key Resistance Level', 'Weekly One Pager']].copy()

        # Rename columns in the copied DataFrame
        selected_columns.rename(columns={
            'Weekly One Pager': 'Game Plan'
        }, inplace=True)

        # Append to Weekly One Pager sheet
        with pd.ExcelWriter(trading_journal_path, mode='a', if_sheet_exists='overlay') as writer:
            worksheet = writer.sheets["Weekly One Pager"]
            start_row = worksheet.max_row  # Find the next empty row
            selected_columns.to_excel(writer, sheet_name="Weekly One Pager", index=False, startrow=start_row, header=False)

        logger.info("weekly_journal_processor: Rows have been copied to the Weekly One Pager sheet successfully.")
        messagebox.showinfo(title="Success", message="Weekly One Pager Built")
    except Exception as e:
        # Log the specific error instead of just the exception object
        logger.error(f"weekly_journal_processor: An error occurred: {e}") 
        # Display error message to user
        messagebox.showerror(title="Error", message=f"An error occurred while processing the journal: {e}")

# ---------------------------- DAILY TASKS --------------------------------- #
def td_import_tasks():
    """Runs the daily tasks as defined in the instructions.
    These are activated via a button in the UI.\n
    - Closes open files so they can be properly handled\n
    - Creates backup of files that will be manipulated\n
    - Imports the tradersync (TDSync) export from the Import folder\n
    - Processes the TDSync export to bring closed trades to the Trade Log\n
    - Processes the Retro data
    - Open file after processing is finished
    #TODO
    """
    close_open_file(trading_journal_path)  # Ensure the file is not open
    close_open_file(tradersync_export_path) # Ensure the file is not open
    create_backup(trading_journal_path)    # Create backup before modifying    
    tradersync_import() # Copy Tradersync export into the TradersyncExport sheet
    new_retro_data_df = process_tradersync_export() # Process the export into the trade log. Only process non OPEN entries.
    if new_retro_data_df is not None:
        process_retro(new_retro_data_df)  # Process the Retrospective data
    open_file(trading_journal_path) # Open the file after processing

def tradersync_import():
    """Copies trade_data.csv content to the TraderSync Export sheet in the trading_journal_path without altering the header."""
    try:
        # Load the CSV file
        trade_data_df = pd.read_csv(tradersync_export_path)

        # Remove dollar signs ('$') from currency columns and convert them to float
        currency_columns = ['Entry Price', 'Exit Price', 'Return $', 'Avg Buy', 'Avg Sell', 
                            'Net Return', 'Commision', 'Strike', 'Cost', 'Fees', 
                            'Return Share', 'Best Exit $']
        trade_data_df[currency_columns] = trade_data_df[currency_columns].replace('[\$,]', '', regex=True).astype(float)

        # Load the Excel file
        book = openpyxl.load_workbook(trading_journal_path)
        sheet = book['TraderSync Export']

        # Save data validations
        data_validations = sheet.data_validations

        # Clear the sheet except the header
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.value = None

        # Apply the currency format to specified columns
        for col_name in currency_columns:
            col_idx = trade_data_df.columns.get_loc(col_name) + 1
            for row_idx, value in enumerate(trade_data_df[col_name], start=2):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.number_format = excel_cell_formater("USD")

        # Write the combined data back to the sheet
        for idx, row in trade_data_df.iterrows():
            for col_idx, value in enumerate(row):
                sheet.cell(row=idx + 2, column=col_idx + 1, value=value)

        # Reapply data validations
        sheet.data_validations = data_validations

        # Save the workbook
        book.save(trading_journal_path)

        logger.info("tradersync_import: Trade data copied to TraderSync Export sheet successfully.")
        messagebox.showinfo(title="Success", message="TraderSync import completed successfully.")
    except Exception as e:
        logger.error(f"tradersync_import: An error occurred: {e}")
        messagebox.showerror(title="Error", message=f"An error occurred while processing the TraderSync import: {e}")

def process_tradersync_export():
    """Reads data from TraderSync Export sheet and appends specific values to Trade Log sheet without altering the headers, only if Status is different from 'OPEN'. Finally, sorts the data by Close Date in descending order and adds a formula to Net Return %."""
    try:
        logger.info("Opening the Excel workbook.")
        # Load the Excel file
        book = openpyxl.load_workbook(trading_journal_path)

        # Load the TraderSync Export and Trade Log sheets
        tradersync_export_sheet = book['TraderSync Export']
        trade_log_sheet = book['Trade Log']

        # Save data validations
        data_validations = trade_log_sheet.data_validations

        # Define the columns to be copied
        columns_to_copy = [
            'Status', 'Symbol', 'Size', 'Open Date', 'Close Date', 'Setups', 'Mistakes', 'Entry Price', 'Exit Price', 'Avg Buy', 
            'Avg Sell', 'Net Return', 'Type', 'MAE', 'MFE', 'Best Exit $', 'Best Exit %'
        ]

        # Get headers from the TraderSync Export sheet
        tradersync_headers = [cell.value for cell in tradersync_export_sheet[1]]

        # Create a dictionary to map column names to indices
        tradersync_col_index = {name: idx + 1 for idx, name in enumerate(tradersync_headers)}

        # Load existing data from Trade Log sheet into a DataFrame
        trade_log_headers = [cell.value for cell in trade_log_sheet[1]]
        existing_data = pd.DataFrame(trade_log_sheet.iter_rows(values_only=True, min_row=2), columns=trade_log_headers)

        # Get the maximum Trade ID from the existing data
        max_trade_id = existing_data['Trade ID'].dropna().astype(int).max() if not existing_data.empty else 0

        # Initialize the next Trade ID
        next_trade_id = max_trade_id + 1

        # Iterate over rows in the TraderSync Export sheet, starting from the second row (skip headers)
        new_data_rows = []
        for row in tradersync_export_sheet.iter_rows(min_row=2, values_only=True):
            logger.info(f"Processing row: {row}")
            # Check if the Status value is different from 'OPEN'
            status_col_index = tradersync_col_index.get('Status')
            if status_col_index and row[status_col_index - 1] != 'OPEN':
                # Create a list to hold the values to be appended
                row_data = [next_trade_id]  # Add the new Trade ID
                for col_name in columns_to_copy:
                    col_index = tradersync_col_index.get(col_name)
                    if col_index:
                        # Handling for Stups and Mistakes. Max 6 Setups and 5 Mistakes.
                        if col_name == 'Setups':
                            setup_counter = 0
                            setups = row[col_index - 1]
                            if isinstance(setups, str) and setups:
                                setup_values = [s.strip() for s in setups.split(',')]
                                for setup_value in setup_values:
                                    row_data.append(setup_value)
                                    setup_counter += 1
                                while setup_counter < 6:
                                    row_data.append(None)
                                    setup_counter += 1
                            else:
                                for _ in range(6):
                                    row_data.append(None)
                        elif col_name == 'Mistakes':
                            mistake_counter = 0
                            mistakes = row[col_index - 1]
                            if isinstance(mistakes, str) and mistakes:
                                mistake_values = [s.strip() for s in mistakes.split(',')]
                                for mistake_value in mistake_values:
                                    row_data.append(mistake_value)
                                    mistake_counter += 1
                                while mistake_counter < 5:
                                    row_data.append(None)
                                    mistake_counter += 1
                            else:
                                for _ in range(5):
                                    row_data.append(None)
                        else:
                            row_data.append(row[col_index - 1])
                    else:
                        row_data.append(None)  # Append None if the column name is not found
                new_data_rows.append(row_data)
                logger.info(f"Appended row data: {row_data}")
                next_trade_id += 1  # Increment the Trade ID for the next entry
        
        logger.debug(f"Passed data: {new_data_rows}")
        # Create a DataFrame for the new data
        new_data_df = pd.DataFrame(new_data_rows, columns=['Trade ID'] + columns_to_copy[:5] + [f'Setup {i}' for i in range(1, 7)] + [f'Mistakes {i}' for i in range(1, 6)]+ columns_to_copy[7:])
        # Combine existing data with the new data
        combined_data = pd.concat([existing_data, new_data_df], ignore_index=True)

        # Sort the combined data by Close Date in descending order. Errors will handle stuff not being recognized as dates.
        combined_data['Close Date'] = pd.to_datetime(combined_data['Close Date'], errors='coerce')
        combined_data = combined_data.sort_values(by='Close Date', ascending=False).reset_index(drop=True)  # Reset index after sorting

        # Clear the Trade Log sheet except the header
        for row in trade_log_sheet.iter_rows(min_row=2, max_row=trade_log_sheet.max_row):
            for cell in row:
                cell.value = None

        # Get column indices for Trade ID, Entry Price, Exit Price, Avg Buy, Avg Sell, Net Return %, Net Return, MAE, MFE, Best Exit $, and Best Exit %
        entry_price_col_idx = trade_log_headers.index('Entry Price') + 1
        exit_price_col_idx = trade_log_headers.index('Exit Price') + 1
        avg_buy_col_idx = trade_log_headers.index('Avg Buy') + 1
        avg_sell_col_idx = trade_log_headers.index('Avg Sell') + 1
        net_return_pct_col_idx = trade_log_headers.index('Net Return %') + 1
        net_return_col_idx = trade_log_headers.index('Net Return') + 1
        mae_col_idx = trade_log_headers.index('MAE') + 1
        mfe_col_idx = trade_log_headers.index('MFE') + 1
        best_exit_col_idx = trade_log_headers.index('Best Exit $') + 1
        best_exit_pct_col_idx = trade_log_headers.index('Best Exit %') + 1

        # Write the sorted data back to the Trade Log sheet and add the formula for Net Return %
        for idx, row in combined_data.iterrows():
            for col_idx, value in enumerate(row, start=1):
                cell = trade_log_sheet.cell(row=idx + 2, column=col_idx, value=value)
                if col_idx in [entry_price_col_idx, exit_price_col_idx, avg_buy_col_idx, avg_sell_col_idx, net_return_col_idx, mae_col_idx, mfe_col_idx, best_exit_col_idx]:
                    cell.number_format = excel_cell_formater("USD")
                elif col_idx == best_exit_pct_col_idx:
                    cell.number_format = excel_cell_formater("%")
                    cell.value = value / 100  # Divide by 100 to correctly show percentage
            
            # Set the formula for Net Return % in the current row
            net_return_pct_cell = trade_log_sheet.cell(row=idx + 2, column=net_return_pct_col_idx)
            net_return_pct_cell.value = f"=({trade_log_sheet.cell(row=idx + 2, column=avg_sell_col_idx).coordinate}-{trade_log_sheet.cell(row=idx + 2, column=avg_buy_col_idx).coordinate})/{trade_log_sheet.cell(row=idx + 2, column=avg_buy_col_idx).coordinate}"
            net_return_pct_cell.number_format = '0.00%'

        logger.info("Finished writing data to Trade Log sheet.")

        # Reapply data validations
        trade_log_sheet.data_validations = data_validations

        # Save the workbook
        book.save(trading_journal_path)
        logger.info("Workbook saved.")

        logger.info("process_tradersync_export: Data from TraderSync Export sheet appended and sorted in Trade Log sheet successfully.")
        messagebox.showinfo(title="Success", message="Data processed, appended, and sorted in Trade Log successfully.")

        return new_data_df
    except Exception as e:
        logger.error(f"process_tradersync_export: An error occurred: {e}")
        messagebox.showerror(title="Error", message=f"An error occurred while processing the TraderSync export: {e}")

def process_retro(retro_df):
    """Appends specific information from the new_data_df to the Retro sheet.
        Args:
        retro_df (pandas dataframe): new pandas data dataframe from daily processing
    """
    try:
        logger.info("Opening the Excel workbook for Retro processing.")
        # Load the Excel file
        book = openpyxl.load_workbook(trading_journal_path)

        # Load the Retro sheet, create if it doesn't exist
        if 'Retro' in book.sheetnames:
            retro_sheet = book['Retro']
        else:
            retro_sheet = book.create_sheet('Retro')

        # Save data validations
        data_validations_retro = retro_sheet.data_validations

        # Load the Data Tags sheet
        data_tags_sheet = book['Data Tags']
        data_tags_df = pd.DataFrame(data_tags_sheet.iter_rows(values_only=True, min_row=2), columns=[cell.value for cell in data_tags_sheet[1]])

        # Save data validations
        data_validations_tags = data_tags_sheet.data_validations

        # Create lists for each column in the Data Tags sheet
        strategy_list = data_tags_df['Strategy'].dropna().tolist()
        sourced_list = data_tags_df['Sourced'].dropna().tolist()
        quality_list = data_tags_df['Quality'].dropna().tolist()
        exit_feeling_list = data_tags_df['Exit Feeling'].dropna().tolist()

        # Create dictionary for current strategies with Retro Due Date
        current_strategies = dict(zip(data_tags_df['Strategy'], data_tags_df['Retro Due Date']))

        # Define the columns to be copied from new_data_df
        columns_to_copy = [
            'Trade ID', 'Symbol', 'Close Date', 'Entry Price', 'Exit Price', 
            'Avg Buy', 'Avg Sell', 'Net Return', 'Status'
        ]

        # Prepare the new data to append
        new_data_to_append = retro_df[columns_to_copy + [f'Setup {i}' for i in range(1, 7)] + [f'Mistakes {i}' for i in range(1, 6)]].copy()

        # Initialize columns for Strategy, Sourced, Quality, Exit Feeling, Retro Due Date
        new_data_to_append.loc[:, 'Strategy'] = None
        new_data_to_append.loc[:, 'Sourced'] = None
        new_data_to_append.loc[:, 'Quality'] = None
        new_data_to_append.loc[:, 'Exit Feeling'] = None
        new_data_to_append.loc[:, 'Retro Due Date'] = None

        # Match Setups with Data Tags lists
        for idx, row in new_data_to_append.iterrows():
            for i in range(1, 7):
                setup_value = row[f'Setup {i}']
                if pd.notna(setup_value):
                    if setup_value in strategy_list:
                        new_data_to_append.loc[idx, 'Strategy'] = setup_value
                        retro_due_date_value = current_strategies.get(setup_value)
                        close_date = pd.to_datetime(row['Close Date'])
                        if retro_due_date_value == "EOD":
                            new_data_to_append.loc[idx, 'Retro Due Date'] = close_date
                        elif retro_due_date_value == "EOW":
                            end_of_week = close_date + timedelta(days=(6 - close_date.weekday()))
                            new_data_to_append.loc[idx, 'Retro Due Date'] = end_of_week
                        elif retro_due_date_value == "D+3":
                            new_data_to_append.loc[idx, 'Retro Due Date'] = close_date + timedelta(days=3)
                    if setup_value in sourced_list:
                        new_data_to_append.loc[idx, 'Sourced'] = setup_value
                    if setup_value in quality_list:
                        new_data_to_append.loc[idx, 'Quality'] = setup_value
            
            for i in range(1, 6):
                mistake_value = row[f'Mistakes {i}']
                if pd.notna(mistake_value) and mistake_value in exit_feeling_list:
                    new_data_to_append.loc[idx, 'Exit Feeling'] = mistake_value

        # Load existing data from Retro sheet into a DataFrame
        retro_headers = [cell.value for cell in retro_sheet[1]]
        existing_data = pd.DataFrame(retro_sheet.iter_rows(values_only=True, min_row=2), columns=retro_headers)

        # Combine existing data with the new data
        combined_data = pd.concat([existing_data, new_data_to_append[columns_to_copy + ['Strategy', 'Sourced', 'Quality', 'Exit Feeling', 'Retro Due Date']]], ignore_index=True)

        # Clear the Retro sheet except the header
        for row in retro_sheet.iter_rows(min_row=2, max_row=retro_sheet.max_row):
            for cell in row:
                cell.value = None

        # Ensure Retro sheet has the correct headers
        for col_idx, header in enumerate(columns_to_copy + ['Strategy', 'Sourced', 'Quality', 'Exit Feeling', 'Retro Due Date'], start=1):
            retro_sheet.cell(row=1, column=col_idx, value=header)

        # Get column indices for Entry Price, Exit Price, Avg Buy, Avg Sell, Net Return
        retro_entry_price_col_idx = retro_headers.index('Entry Price') + 1
        retro_exit_price_col_idx = retro_headers.index('Exit Price') + 1
        retro_avg_buy_col_idx = retro_headers.index('Avg Buy') + 1
        retro_avg_sell_col_idx = retro_headers.index('Avg Sell') + 1
        retro_net_return = retro_headers.index('Net Return') + 1
        retro_retro_due_date = retro_headers.index('Retro Due Date') + 1

        # Write the combined data back to the Retro sheet, and apply formats to desired cells
        for idx, row in combined_data.iterrows():
            for col_idx, value in enumerate(row, start=1):
                retro_cell = retro_sheet.cell(row=idx + 2, column=col_idx, value=value)
                if col_idx in [retro_entry_price_col_idx, retro_exit_price_col_idx, retro_avg_buy_col_idx, retro_avg_sell_col_idx, retro_net_return]:
                    retro_cell.number_format = excel_cell_formater("USD")
                elif col_idx in [retro_retro_due_date]:
                    retro_cell.number_format = excel_cell_formater("date")

        logger.info("Finished writing data to Retro sheet.")

        # Reapply data validations
        retro_sheet.data_validations = data_validations_retro
        data_tags_sheet.data_validations = data_validations_tags

        # Save the workbook
        book.save(trading_journal_path)
        logger.info("Workbook saved.")

        logger.info("process_retro: Data from new_data_df appended to Retro sheet successfully.")
        messagebox.showinfo(title="Success", message="Data processed and appended to Retro sheet successfully.")
    except Exception as e:
        logger.error(f"process_retro: An error occurred: {e}")
        messagebox.showerror(title="Error", message=f"An error occurred while processing the Retro data: {e}")

# ---------------------------- STRATEGY DB --------------------------------- #
def update_strategies():
    """Runs the strategy DB update task.
    This will create a dictionary with all the current strategies and related info.
    """
    pass

# ---------------------------- REMINDERS ----------------------------------- #
def update_reminders_tasks():
    """Runs the update reminders tasks as indicated in instructions.
    Goes through the Journal, Ideas and Retro sheets to detect which are missing
    information like retro results and Trade IDs.
    """
    close_open_file(trading_journal_path)  # Ensure the file is not open
    update_reminders()  # Goes through the file and updates reminders
    open_file(trading_journal_path) # Open the file after processing

def update_reminders():
    """Runs the update reminders task.
    This will go through all trades and make sure to highlight anything that needs to be retroed.
    """
    try:
        logger.info("Opening the Excel workbook for updating reminders.")
        # Load the Excel file
        book = openpyxl.load_workbook(trading_journal_path)

        # Load the Journal sheet
        journal_sheet = book['Journal']
        journal_df = pd.DataFrame(journal_sheet.iter_rows(values_only=True, min_row=2), columns=[cell.value for cell in journal_sheet[1]])

        # Check for entries older than 30 days and mark as 'DUE' in the '30D Retro' column
        today = pd.to_datetime(datetime.now().date())
        journal_df['Date'] = pd.to_datetime(journal_df['Date'])
        journal_df.loc[(today - journal_df['Date']).dt.days > 30, '30D Retro'] = 'DUE'

        # Update the Journal sheet
        for idx, row in journal_df.iterrows():
            for col_idx, value in enumerate(row, start=1):
                journal_sheet.cell(row=idx + 2, column=col_idx, value=value)

        # Load the Retro sheet
        retro_sheet = book['Retro']
        retro_df = pd.DataFrame(retro_sheet.iter_rows(values_only=True, min_row=2), columns=[cell.value for cell in retro_sheet[1]])

        # Check for Retro Due Date entries that have passed and mark as 'DUE' in the 'Retro Due Date' column
        retro_due_dates = []
        for idx, row in retro_df.iterrows():
            due_date = row['Retro Due Date']
            if due_date != 'DUE':
                if due_date is not None and pd.to_datetime(due_date, errors='coerce') < today:
                    retro_due_dates.append('DUE')
                else:
                    retro_due_dates.append(due_date)
            else:
                retro_due_dates.append(due_date)
        retro_df['Retro Due Date'] = retro_due_dates

        # Update the Retro sheet
        for idx, row in retro_df.iterrows():
            for col_idx, value in enumerate(row, start=1):
                retro_sheet.cell(row=idx + 2, column=col_idx, value=value)
                
        # Load the Ideas sheet
        ideas_sheet = book['Ideas']
        ideas_df = pd.DataFrame(ideas_sheet.iter_rows(values_only=True, min_row=2), columns=[cell.value for cell in ideas_sheet[1]])

        # Check for Date entries older than 1 day and update the Filter and Trade ID columns
        ideas_df['Date'] = pd.to_datetime(ideas_df['Date'], errors='coerce')
        expired_mask = (today - ideas_df['Date']).dt.days > 30

        # Update Filter column to "Expired" if older than 30 days and Filter is empty
        ideas_df.loc[expired_mask & ideas_df['Filter'].isna(), 'Filter'] = 'Expired'

        # Add "DUE" to Trade ID if older than 1 day, Filter is not "Expired", and Trade ID is empty
        ideas_df.loc[(ideas_df['Filter'] == 'Taken') & ideas_df['Trade ID'].isna(), 'Trade ID'] = 'DUE'

        # Update the Ideas sheet
        for idx, row in ideas_df.iterrows():
            for col_idx, value in enumerate(row, start=1):
                ideas_sheet.cell(row=idx + 2, column=col_idx, value=value)

        # Save the workbook
        book.save(trading_journal_path)
        logger.info("update_reminders: Reminders updated successfully.")
        messagebox.showinfo(title="Success", message="Reminders updated successfully.")
    except Exception as e:
        logger.error(f"update_reminders: An error occurred: {e}")
        messagebox.showerror(title="Error", message=f"An error occurred while updating reminders: {e}")

# ---------------------------- UI SETUP ------------------------------------ #
# Main window UI setup
main_window = Tk()
main_window.title("Trading Journal")
main_window.config(padx=50, pady=50, bg=WHITE)

# Action button that will run the weekly tasks
weekly_task_button = Button(text="Process Journal", command=process_one_pager_task)
weekly_task_button.grid(column=0, row=0)

# Action button that will run the daily tasks
td_sync_button = Button(text="Perform Tradersync Import", command=td_import_tasks)
td_sync_button.grid(column=0, row=1)

# Action button that will update reminders
update_reminders_button = Button(text="Update Reminders", command=update_reminders_tasks)
update_reminders_button.grid(column=0, row=2)

# Action button that will run the strategy update task
strategy_update_button = Button(text="Update Strategies WIP", command=update_strategies)
strategy_update_button.grid(column=0, row=3)

main_window.mainloop()
